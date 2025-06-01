# handlers/admin.py
from aiogram import Router, F
from aiogram.filters import Command
from aiogram.types import Message, CallbackQuery, Document
from aiogram.utils.keyboard import InlineKeyboardBuilder
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State
from database.db import get_db_session, User, Application, Event, AllowedUser, get_or_create_group, Group, Semester
from sqlalchemy.exc import IntegrityError
import openpyxl
import io
import os
import openpyxl
from datetime import datetime
from aiogram.types import ContentType, Message
from aiogram.filters import StateFilter
from aiogram.fsm.context import FSMContext

from config import ADMIN_IDS

router = Router()

class UploadExcel(StatesGroup):
    type = State()

class FindStudent(StatesGroup):
    query = State()

class SetRole(StatesGroup):
    waiting_for_telegram_id = State()
    waiting_for_role = State()

# =============================
# 1. Главное админ-меню
# =============================
async def show_admin_menu(message: Message):
    """Показывает клавиатуру администратора."""
    kb = InlineKeyboardBuilder()
    kb.button(text="📋 Пользователи", callback_data="admin_users")
    kb.button(text="🔍 Поиск студента", callback_data="admin_find_user")
    kb.button(text="📊 Отчёты", callback_data="admin_stats")
    kb.button(text="🧹 Очистить заявки", callback_data="admin_clear_apps")
    kb.button(text="📅 Импорт расписания (Excel)", callback_data="admin_upload_schedule")
    kb.button(text="📤 Импорт списка студентов (Excel)", callback_data="admin_upload_excel")
    
    kb.adjust(1)
    await message.answer("🛠 <b>Админ-панель</b>", reply_markup=kb.as_markup())


# =============================
# 2. Вход в админ-панель
# =============================
@router.message(Command("admin"))
async def admin_panel_cmd(message: Message):
    if message.from_user.id not in ADMIN_IDS:
        await message.answer("❌ У вас нет прав администратора.")
        return
    await show_admin_menu(message)


# =============================
# 3. Просмотр списка пользователей
# =============================
@router.callback_query(F.data == "admin_users")
async def admin_users(callback: CallbackQuery):
    session = get_db_session()
    users = session.query(User).order_by(User.id).all()
    if not users:
        await callback.answer("Пользователи не найдены.")
        session.close()
        return

    for user in users[:50]:
        kb = InlineKeyboardBuilder()
        kb.button(text="❌ Удалить", callback_data=f"admin_delete_user_{user.id}")
        await callback.message.answer(
            f"👤 <b>{user.full_name}</b> — <a href='tg://user?id={user.telegram_id}'>[написать]</a>\n"
            f"🏫 Группа: {user.group.name if user.group else '—'}\n"
            f"🆔 Telegram ID: <code>{user.telegram_id}</code>",
            reply_markup=kb.as_markup()
        )
    await callback.answer("✅ Список пользователей отправлен.")
    session.close()

@router.callback_query(F.data == "admin_find_user")
async def admin_find_user(callback: CallbackQuery, state: FSMContext):
    await callback.message.edit_text("🔎 Введите ФИО, группу или Telegram ID для поиска:")
    await state.set_state(FindStudent.query)

@router.callback_query(F.data == "admin_upload_schedule")
async def prompt_schedule_upload(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("📎 Отправьте Excel-файл с расписанием занятий.\n"
                                  "Формат: Группа | День недели | Время | Предмет | Преподаватель | Аудитория | Неделя")
    await state.set_state(UploadExcel.type)
    await state.update_data(file_type="schedule")
    await callback.answer()


@router.message(FindStudent.query)
async def process_find_student(message: Message, state: FSMContext):
    text = message.text.strip()
    session = get_db_session()

    results = session.query(User).filter(
        (User.full_name.ilike(f"%{text}%")) |
        (User.telegram_id == text if text.isdigit() else False) |
        (User.group.has(name=text.upper()))
    ).all()

    if not results:
        await message.answer("❌ Пользователь не найден.")
    else:
        for user in results[:10]:  # ограничим до 10
            kb = InlineKeyboardBuilder()
            kb.button(text="❌ Удалить", callback_data=f"admin_delete_user_{user.id}")
            await message.answer(
                f"👤 <b>{user.full_name}</b> — <a href='tg://user?id={user.telegram_id}'>[написать]</a>\n"
                f"🏫 Группа: {user.group.name if user.group else '—'}\n"
                f"🆔 Telegram ID: <code>{user.telegram_id}</code>",
                reply_markup=kb.as_markup()
            )

    await state.clear()
    session.close()


# =============================
# 4. Удаление пользователя вручную
# =============================
@router.callback_query(F.data.startswith("admin_delete_user_"))
async def admin_delete_user(callback: CallbackQuery):
    user_id = int(callback.data.split("_")[-1])
    session = get_db_session()
    user = session.query(User).filter_by(id=user_id).first()
    if user:
        # Удаляем связанные заявки перед удалением пользователя
        session.query(Application).filter_by(user_id=user.id).delete()
        session.delete(user)
        session.commit()
        await callback.answer("✅ Пользователь удалён.", show_alert=True)
        await callback.message.edit_text("❌ Пользователь удалён.")
    else:
        await callback.answer("❌ Пользователь не найден.")
    session.close()


# =============================
# 5. Статистика проекта
# =============================
@router.callback_query(F.data == "admin_stats")
async def admin_stats(callback: CallbackQuery):
    session = get_db_session()
    total_users = session.query(User).count()
    total_apps = session.query(Application).count()
    total_events = session.query(Event).count()
    active_events = session.query(Event).filter_by(is_active=1).count()

    text = (
        "📊 <b>Статистика проекта</b>\n\n"
        f"👨‍🎓 Зарегистрировано студентов: <b>{total_users}</b>\n"
        f"✉ Подано заявок: <b>{total_apps}</b>\n"
        f"🎉 Всего мероприятий: <b>{total_events}</b>\n"
        f"🟢 Активных мероприятий: <b>{active_events}</b>"
    )
    await callback.message.answer(text)
    await callback.answer()
    session.close()


# =============================
# 6. Очистка всех заявок
# =============================
@router.callback_query(F.data == "admin_clear_apps")
async def admin_clear_apps(callback: CallbackQuery):
    kb = InlineKeyboardBuilder()
    kb.button(text="✅ Подтвердить", callback_data="admin_clear_confirm")
    kb.button(text="❌ Отмена", callback_data="admin_clear_cancel")
    kb.adjust(2)
    await callback.message.answer("⚠ Вы действительно хотите удалить <b>ВСЕ</b> заявки?", reply_markup=kb.as_markup())
    await callback.answer()

@router.callback_query(F.data == "admin_clear_cancel")
async def admin_clear_cancel(callback: CallbackQuery):
    await callback.answer("Очистка отменена.", show_alert=False)
    await show_admin_menu(callback.message)

@router.callback_query(F.data == "admin_clear_confirm")
async def admin_clear_confirm(callback: CallbackQuery):
    session = get_db_session()
    deleted = session.query(Application).delete()
    session.commit()
    session.close()
    await callback.answer(f"✅ Удалено заявок: {deleted}", show_alert=True)
    await show_admin_menu(callback.message)



# Сброс всех FSM состояний /admin_reset_all_fsm

@router.message(Command("admin_reset_all_fsm"))
async def admin_reset_all_fsm(message: Message, state: FSMContext):
    if message.from_user.id not in ADMIN_IDS:
        await message.answer("❌ У вас нет прав администратора.")
        return

    await state.clear()
    await message.answer("✅ Все FSM-состояния текущего пользователя сброшены.")

# Назначение роли /set_role
@router.message(Command("set_role"))
async def cmd_set_role(message: Message, state: FSMContext):
    await message.answer("Введите Telegram ID пользователя, которому нужно назначить роль:")
    await state.set_state(SetRole.waiting_for_telegram_id)

@router.message(SetRole.waiting_for_telegram_id)
async def process_telegram_id(message: Message, state: FSMContext):
    telegram_id = message.text.strip()
    if not telegram_id.isdigit():
        await message.answer("❌ Введите корректный числовой Telegram ID.")
        return
    await state.update_data(telegram_id=int(telegram_id))
    await message.answer("Введите новую роль для пользователя: student / dean / admin")
    await state.set_state(SetRole.waiting_for_role)

@router.message(SetRole.waiting_for_role)
async def process_new_role(message: Message, state: FSMContext):
    role = message.text.strip().lower()
    if role not in ["student", "dean", "admin"]:
        await message.answer("❌ Недопустимая роль. Введите: student / dean / admin")
        return

    data = await state.get_data()
    telegram_id = data["telegram_id"]

    session = get_db_session()
    user = session.query(User).filter_by(telegram_id=telegram_id).first()
    if not user:
        await message.answer("❌ Пользователь не найден.")
    else:
        user.role = role
        session.commit()
        await message.answer(f"✅ Роль пользователя <b>{user.full_name}</b> обновлена на <b>{role}</b>.")

    await state.clear()
    session.close()

# Обработчик кнопки загрузки excel
@router.callback_query(F.data == "admin_upload_excel")
async def prompt_excel_upload(callback: CallbackQuery, state: FSMContext):
    await callback.message.answer("📎 Отправьте Excel-файл (.xlsx) со списком студентов.\n"
                                  "Формат: <code>ФИО | Группа</code> (первая строка — заголовки).")
    await state.set_state(UploadExcel.type)
    await state.update_data(file_type="students")
    await callback.answer()

#Обработка excel
@router.message(F.content_type == ContentType.DOCUMENT, StateFilter(UploadExcel.type))
async def handle_excel_file(message: Message, state: FSMContext):
    os.makedirs("temp", exist_ok=True) # Убедимся, что папка temp существует
    document = message.document
    file_path = f"temp/{document.file_name}"
    await message.bot.download(document, destination=file_path)

    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    session = get_db_session()
    added = 0

    file_type = (await state.get_data()).get("file_type")

    if file_type == "schedule":
        from database.db import Schedule, get_or_create_group, Semester, Group
        from datetime import datetime

        semester_dates = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            sem_group = str(row[8]).strip() if row[8] else None
            start_date = row[9]
            end_date = row[10]
            if sem_group and isinstance(start_date, datetime) and isinstance(end_date, datetime):
                semester_dates[sem_group] = (start_date.date(), end_date.date())

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not all(row[:7]):
                continue

            print(f"Импортируем строку: {row}")

            group_name = str(row[0]).strip()
            day = str(row[1]).strip()
            time = str(row[2]).strip()
            subject = str(row[3]).strip()
            teacher = str(row[4]).strip()
            room = str(row[5]).strip()
            week = int(str(row[6]).strip()) if str(row[6]).strip() in ['1', '2'] else 1

            group = get_or_create_group(session, group_name)

            if group_name in semester_dates:
                start_date, end_date = semester_dates[group_name]

                existing_semester = (
                    session.query(Semester)
                    .join(Semester.groups)
                    .filter(Group.id == group.id)
                    .first()
                )
                if not existing_semester:
                    semester = Semester(
                        number=1,
                        date_start=start_date,
                        date_end=end_date,
                        group_name=group_name,
                        groups=[group]
                    )
                    session.add(semester)

            new_schedule = Schedule(
                group_id=group.id,
                day_of_week=day.upper(),
                time=time,
                subject=subject,
                teacher=teacher,
                room=room,
                week_number=week
            )
            session.add(new_schedule)
            added += 1

        await message.answer(f"✅ Импорт расписания завершён. Добавлено {added} записей.")

    session.commit()
    session.close()
    await state.clear()




# Регистрация роутера

def register(dp):
    dp.include_router(router)